"""报表导出服务：HTML / Excel 导出。"""

from __future__ import annotations
import json
from io import BytesIO
from datetime import datetime

import jinja2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backend.services.ticket_processor import TicketProcessor


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }}</title>
<style>
  :root {
    --bg: #0f1923; --bg-elevated: #152238; --card-bg: #1a2837; --card-border: #2a3f54;
    --text: #e0e8f0; --text-secondary: #8899aa;
    --accent: #00d4ff; --accent2: #ff6b6b; --accent3: #ffd93d; --accent4: #6bcb77;
  }
  * { margin:0; padding:0; box-sizing:border-box; }
  body { background:var(--bg); color:var(--text); font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif; line-height:1.6; }

  /* Header */
  .header {
    background:linear-gradient(135deg,#0a1628 0%,#152238 50%,#0d2137 100%);
    padding:32px 40px; border-bottom:1px solid var(--card-border);
    position:sticky; top:0; z-index:100;
    backdrop-filter:blur(12px);
  }
  .header h1 { font-size:28px; font-weight:700; }
  .header h1 span { color:var(--accent); }
  .header .subtitle { color:var(--text-secondary); font-size:14px; margin-top:6px; }

  /* Layout */
  .layout { display:flex; min-height:calc(100vh - 120px); }

  /* TOC */
  .toc {
    width:220px; background:var(--bg-elevated); border-right:1px solid var(--card-border);
    padding:20px 16px; position:sticky; top:120px; height:calc(100vh - 120px);
    overflow-y:auto; flex-shrink:0;
  }
  .toc h3 { font-size:13px; color:var(--text-secondary); text-transform:uppercase; letter-spacing:1px; margin-bottom:12px; }
  .toc a {
    display:block; padding:6px 10px; font-size:13px; color:var(--text-secondary);
    border-radius:6px; text-decoration:none; transition:all 0.2s;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
  }
  .toc a:hover { background:rgba(0,212,255,0.08); color:var(--accent); }
  .toc a.active { background:rgba(0,212,255,0.12); color:var(--accent); font-weight:600; }

  /* Content */
  .content { flex:1; padding:28px 40px; }

  /* Chart cards */
  .chart-card {
    background:var(--card-bg); border:1px solid var(--card-border); border-radius:12px;
    padding:20px; margin-bottom:20px; break-inside:avoid;
    transition:border-color 0.2s;
  }
  .chart-card:hover { border-color:rgba(0,212,255,0.3); }
  .chart-title { font-size:15px; font-weight:600; margin-bottom:4px; }
  .chart-desc { font-size:12px; color:var(--text-secondary); margin-bottom:12px; }

  /* Insights */
  .insight-item { padding:14px 18px; margin-bottom:12px; border-radius:8px; border-left:4px solid; background:rgba(255,255,255,0.02); }
  .insight-item.critical { border-color:var(--accent2); background:rgba(255,107,107,0.06); }
  .insight-item.warning { border-color:var(--accent3); background:rgba(255,217,61,0.06); }
  .insight-item.info { border-color:var(--accent); background:rgba(0,212,255,0.06); }
  .insight-item .tag { display:inline-block; font-size:11px; padding:2px 8px; border-radius:4px; font-weight:600; margin-right:8px; }
  .critical .tag { background:rgba(255,107,107,0.2); color:var(--accent2); }
  .warning .tag { background:rgba(255,217,61,0.2); color:var(--accent3); }
  .info .tag { background:rgba(0,212,255,0.2); color:var(--accent); }
  .insight-item .title { font-weight:600; font-size:14px; }
  .insight-item .desc { font-size:13px; color:var(--text-secondary); margin-top:4px; }

  /* Table */
  table { width:100%; border-collapse:collapse; margin-top:16px; }
  th, td { padding:8px 12px; text-align:left; border-bottom:1px solid var(--card-border); }
  th { color:var(--text-secondary); font-size:12px; font-weight:600; background:var(--bg-elevated); }
  tbody tr:hover { background:rgba(0,212,255,0.03); }

  /* Footer */
  .footer { text-align:center; padding:24px; color:var(--text-secondary); font-size:12px; border-top:1px solid var(--card-border); }

  /* Scroll to top */
  .scroll-top {
    position:fixed; bottom:24px; right:24px; width:40px; height:40px;
    background:var(--gradient1); color:white; border:none; border-radius:50%;
    cursor:pointer; font-size:18px; display:none; align-items:center; justify-content:center;
    box-shadow:0 4px 16px rgba(0,212,255,0.3); z-index:200;
  }
  .scroll-top.visible { display:flex; }

  /* Print */
  @media print {
    body { background:white; color:#333; }
    .header { position:static; background:white; color:#333; border-bottom:2px solid #333; }
    .header h1 span { color:#0078d4; }
    .toc { display:none; }
    .scroll-top { display:none !important; }
    .chart-card { break-inside:avoid; border:1px solid #ddd; }
    .footer { border-top:1px solid #ddd; }
  }

  /* Responsive */
  @media (max-width:768px) {
    .layout { flex-direction:column; }
    .toc { width:100%; position:static; height:auto; border-right:none; border-bottom:1px solid var(--card-border); }
    .header { padding:20px; }
    .content { padding:20px; }
    .chart-card { padding:14px; }
  }
</style>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
</head>
<body>
<div class="header">
  <h1>{{ title }}<span>报告</span></h1>
  <div class="subtitle">数据来源：工单分析 | 生成日期：{{ generated_at }} | 总计 {{ total }} 件工单</div>
</div>
<div class="layout">
<nav class="toc">
  <h3>目录</h3>
  {% for chart in charts %}
  <a href="#{{ chart.id }}">{{ chart.title }}</a>
  {% endfor %}
  {% if insights %}
  <a href="#insights">关键洞察</a>
  {% endif %}
  {% if data_table %}
  <a href="#data-table">数据明细</a>
  {% endif %}
</nav>
<div class="content">
{% for chart in charts %}
<div class="chart-card" id="{{ chart.id }}">
  <div class="chart-title">{{ chart.title }}</div>
  <div id="chart_{{ chart.id }}" style="width:100%;height:350px;"></div>
</div>
{% endfor %}

{% if insights %}
<div class="chart-card" id="insights">
  <div class="chart-title">关键洞察</div>
  {% for insight in insights %}
  <div class="insight-item {{ insight.severity }}">
    <span class="tag">{{ insight.severity | severity_label }}</span>
    <span class="title">{{ insight.title }}</span>
    <div class="desc">{{ insight.desc }}</div>
  </div>
  {% endfor %}
</div>
{% endif %}

{% if data_table %}
<div class="chart-card" id="data-table">
  <div class="chart-title">数据明细</div>
  <table>
    <thead><tr>{% for h in data_table.headers %}<th>{{ h }}</th>{% endfor %}</tr></thead>
    <tbody>{% for row in data_table.rows %}<tr>{% for cell in row %}<td>{{ cell }}</td>{% endfor %}</tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}
</div>
</div>
<div class="footer">{{ title }} | 智能报表分析 Agent 平台 | {{ generated_at }}</div>
<button class="scroll-top" id="scrollTop" onclick="window.scrollTo({top:0,behavior:'smooth'})">&#x2191;</button>
<script>
var charts = {{ charts_json | safe }};
charts.forEach(function(c) {
  var dom = document.getElementById('chart_' + c.id);
  if (dom) { var chart = echarts.init(dom, null, { renderer: 'canvas' }); chart.setOption(c.option); }
});
window.addEventListener('resize', function() { charts.forEach(function(c) { var d=document.getElementById('chart_'+c.id); if(d) echarts.getInstanceByDom(d)?.resize(); }); });

// TOC active tracking
var tocLinks = document.querySelectorAll('.toc a');
var observer = new IntersectionObserver(function(entries) {
  entries.forEach(function(e) {
    if (e.isIntersecting) {
      tocLinks.forEach(function(l) { l.classList.remove('active'); });
      var link = document.querySelector('.toc a[href="#' + e.target.id + '"]');
      if (link) link.classList.add('active');
    }
  });
}, { rootMargin: '-20% 0px -60% 0px' });
document.querySelectorAll('.chart-card, #insights, #data-table').forEach(function(el) { observer.observe(el); });

// Scroll to top button
window.addEventListener('scroll', function() {
  var btn = document.getElementById('scrollTop');
  if (window.scrollY > 300) btn.classList.add('visible'); else btn.classList.remove('visible');
});
</script>
</body>
</html>
"""

SEVERITY_LABELS = {'critical': '严重', 'warning': '警告', 'info': '关注'}

# Excel style presets
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=12)
HEADER_FILL = PatternFill(start_color='1a2837', end_color='1a2837', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
BODY_FONT = Font(name='Microsoft YaHei', size=11)
BODY_ALIGN = Alignment(horizontal='left', vertical='center')
ALT_ROW_FILL = PatternFill(start_color='152238', end_color='152238', fill_type='solid')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=24, color='00d4ff')
SUBTITLE_FONT = Font(name='Microsoft YaHei', size=12, color='8899aa')
KPI_FONT = Font(name='Microsoft YaHei', bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style='thin', color='2a3f54'),
    right=Side(style='thin', color='2a3f54'),
    top=Side(style='thin', color='2a3f54'),
    bottom=Side(style='thin', color='2a3f54'),
)
SEVERITY_FILLS = {
    'critical': PatternFill(start_color='442222', end_color='442222', fill_type='solid'),
    'warning': PatternFill(start_color='443311', end_color='443311', fill_type='solid'),
    'info': PatternFill(start_color='112244', end_color='112244', fill_type='solid'),
}


def _style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data_rows(ws, num_rows, num_cols):
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL


def export_html(
    title: str,
    charts: list[dict],
    insights: list[dict],
    data_table: dict | None = None,
    total: int = 0,
) -> bytes:
    """生成独立 HTML 报告文件。"""
    env = jinja2.Environment()
    env.filters['severity_label'] = lambda x: SEVERITY_LABELS.get(x, x)

    template = env.from_string(HTML_TEMPLATE)
    charts_json = json.dumps([
        {'id': c['id'], 'title': c['title'], 'option': c['option']}
        for c in charts
    ], ensure_ascii=False)

    html = template.render(
        title=title,
        generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'),
        total=total,
        charts=charts,
        charts_json=charts_json,
        insights=insights,
        data_table=data_table,
    )
    return html.encode('utf-8')


def export_excel(
    title: str,
    data_table: dict,
    kpis: dict,
    insights: list[dict],
    charts: list[dict] | None = None,
) -> bytes:
    """生成 Excel 报告（封面 + 数据表 + KPI + 洞察 + 图表数据）。"""
    wb = Workbook()

    # ── Sheet 0: 封面 ──────────────────────────────────
    ws0 = wb.active
    ws0.title = '封面'
    ws0.merge_cells('A1:B1')
    title_cell = ws0['A1']
    title_cell.value = title
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws0.row_dimensions[1].height = 50

    ws0.merge_cells('A2:B2')
    sub_cell = ws0['A2']
    sub_cell.value = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    sub_cell.font = SUBTITLE_FONT

    # KPI summary table on cover
    ws0.merge_cells('A4:B4')
    ws0['A4'].value = '关键指标'
    ws0['A4'].font = Font(name='Microsoft YaHei', bold=True, size=14, color='e0e8f0')

    kpi_items = [
        ('总工单数', kpis.get('total', 0)),
        ('SLA达标率', f'{kpis.get("sla_ratio", 0)}%'),
        ('平均解决时间', f'{kpis.get("avg_resolution_days", 0)}天'),
    ]
    for i, (label, val) in enumerate(kpi_items, start=5):
        ws0.cell(row=i, column=1, value=label).font = BODY_FONT
        ws0.cell(row=i, column=2, value=val).font = KPI_FONT

    ws0.column_dimensions['A'].width = 16
    ws0.column_dimensions['B'].width = 36

    # ── Sheet 1: 数据明细 ───────────────────────────────
    ws1 = wb.create_sheet('数据明细')
    if data_table and data_table.get('headers'):
        ws1.append(data_table['headers'])
        for row in data_table.get('rows', []):
            ws1.append(row)
        _style_header_row(ws1, len(data_table['headers']))
        _style_data_rows(ws1, len(data_table['rows']) + 1, len(data_table['headers']))
        # Auto-filter
        ws1.auto_filter.ref = f'A1:{chr(64 + len(data_table["headers"]))}{len(data_table["rows"]) + 1}'
        # Freeze panes
        ws1.freeze_panes = 'A2'

    # ── Sheet 2: KPI汇总 ───────────────────────────────
    ws2 = wb.create_sheet('KPI汇总')
    ws2.append(['指标', '值'])
    ws2.append(['总工单数', kpis.get('total', 0)])
    ws2.append(['SLA达标率', f'{kpis.get("sla_ratio", 0)}%'])
    ws2.append(['平均解决时间', f'{kpis.get("avg_resolution_days", 0)}天'])
    _style_header_row(ws2, 2)
    _style_data_rows(ws2, 7, 2)
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 36

    # ── Sheet 3: 洞察建议 ──────────────────────────────
    ws3 = wb.create_sheet('洞察建议')
    ws3.append(['级别', '标题', '描述'])
    for insight in insights:
        ws3.append([
            SEVERITY_LABELS.get(insight.get('severity', ''), insight.get('severity', '')),
            insight.get('title', ''),
            insight.get('desc', ''),
        ])
    _style_header_row(ws3, 3)
    num_rows3 = len(insights) + 1
    for row in range(2, num_rows3 + 1):
        for col in range(1, 4):
            cell = ws3.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            # Color by severity
            sev = insights[row - 2].get('severity', '') if row - 2 < len(insights) else ''
            if sev in SEVERITY_FILLS:
                cell.fill = SEVERITY_FILLS[sev]
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 50

    # ── Sheet 4: 图表数据 (optional) ───────────────────
    if charts:
        ws4 = wb.create_sheet('图表数据')
        ws4.append(['图表名称', '标签', '数值'])
        _style_header_row(ws4, 3)
        r = 2
        for chart in charts:
            opt = chart.get('option', {})
            for series in opt.get('series', []):
                sname = series.get('name', chart.get('title', ''))
                for item in series.get('data', []):
                    if isinstance(item, dict):
                        ws4.append([sname, item.get('name', ''), item.get('value', '')])
                    else:
                        ws4.append([sname, '', item])
                    _style_data_rows(ws4, r + 1, 3)
                    r += 1
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 24
        ws4.column_dimensions['C'].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
